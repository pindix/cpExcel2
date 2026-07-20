import pandas as pd
from openpyxl import load_workbook

def motor_horizontal(arquivo_nome, lista_medicamentos):
    try:
        # 1. Abre o ficheiro preservando estilos e cabeçalhos
        wb = load_workbook(arquivo_nome)
        ws = wb['oms']

        # 2. Localizar a primeira linha vaga a partir da 8
        linha_vaga = 8
        while ws.cell(row=linha_vaga, column=1).value is not None:
            linha_vaga += 1
        
        print(f"Começando a inserir dados a partir da linha: {linha_vaga}")

        # 3. Inserção cega (Esquerda para Direita)
        # Cada 'med' deve ter exatamente 9 itens para bater com as colunas A até I
        for med in lista_medicamentos:
            for col_idx, valor in enumerate(med, start=1):
                ws.cell(row=linha_vaga, column=col_idx).value = valor
            linha_vaga += 1

        # 4. Salvar no próprio arquivo
        wb.save(arquivo_nome)
        print(f"Sucesso! {len(lista_medicamentos)} linhas adicionadas.")

    except Exception as e:
        print(f"Erro no Pydroid: {e}")

# --- LISTA DE DADOS (Exemplo com as tuas novas regras de campos vazios) ---
# Estrutura: [nome, doenca, dose, via, campos, concentracao, intervalo, formula, nota]
novos_dados = [
    ["Antimoniato de meglumina", "Leishmaniose", "adulta", "IM", "peso(40, 150); Idade(6570, 36500) dosagem(20, 20, 20, mg/kg).", "81", "24", "", ""],
    ["Antimoniato de meglumina", "Leishmaniose", "pediatrica", "IM", "peso(1, 40); Idade(30, 6570) dosagem(20, 20, 20, mg/kg).", "81", "24", "", ""],
    ["Antimoniato de meglumina", "Leishmaniose", "adulta", "IV", "peso(40, 150); Idade(6570, 36500) dosagem(20, 20, 20, mg/kg).", "81", "24", "", ""],
    ["Antimoniato de meglumina", "Leishmaniose", "pediatrica", "IV", "peso(1, 40); Idade(30, 6570) dosagem(20, 20, 20, mg/kg).", "81", "24", "", ""],
    ["Estibogliconato de sodio", "Leishmaniose", "adulta", "IM", "peso(40, 150); Idade(6570, 36500) dosagem(20, 20, 20, mg/kg).", "100", "24", "", ""],
    ["Estibogliconato de sodio", "Leishmaniose", "pediatrica", "IM", "peso(1, 40); Idade(30, 6570) dosagem(20, 20, 20, mg/kg).", "100", "24", "", ""],
    ["Estibogliconato de sodio", "Leishmaniose", "adulta", "IV", "peso(40, 150); Idade(6570, 36500) dosagem(20, 20, 20, mg/kg).", "100", "24", "", ""],
    ["Estibogliconato de sodio", "Leishmaniose", "pediatrica", "IV", "peso(1, 40); Idade(30, 6570) dosagem(20, 20, 20, mg/kg).", "100", "24", "", ""],
    ["Paromomicina", "Leishmaniose visceral", "adulta", "IM", "peso(40, 150); Idade(6570, 36500) dosagem(15, 15, 15, mg/kg).", "750", "24", "", ""],
    ["Paromomicina", "Leishmaniose visceral", "pediatrica", "IM", "peso(1, 40); Idade(30, 6570) dosagem(15, 15, 15, mg/kg).", "750", "24", "", ""],
    ["Pentamidina", "Tripanossomíase africana", "adulta", "IM", "peso(40, 150); Idade(6570, 36500) dosagem(4, 4, 4, mg/kg).", "200, 300", "24", "", ""],
    ["Pentamidina", "Tripanossomíase africana", "pediatrica", "IM", "peso(1, 40); Idade(30, 6570) dosagem(4, 4, 4, mg/kg).", "200, 300", "24", "", ""],
    ["Pentamidina", "Tripanossomíase africana", "adulta", "IV", "peso(40, 150); Idade(6570, 36500) dosagem(4, 4, 4, mg/kg).", "200, 300", "24", "", ""],
    ["Pentamidina", "Tripanossomíase africana", "pediatrica", "IV", "peso(1, 40); Idade(30, 6570) dosagem(4, 4, 4, mg/kg).", "200, 300", "24", "", ""],
    ["Suramina sodica", "Tripanossomíase africana", "adulta", "IV", "peso(40, 150); Idade(6570, 36500) dosagem(20, 20, 20, mg/kg).", "1000", "168", "", ""],
    ["Suramina sodica", "Tripanossomíase africana", "pediatrica", "IV", "peso(1, 40); Idade(30, 6570) dosagem(20, 20, 20, mg/kg).", "1000", "168", "", ""],
    ["Melarsoprol", "Tripanossomíase africana", "adulta", "IV", "peso(40, 150); Idade(6570, 36500) dosagem(2.2, 3.6, 2.2, mg/kg).", "36", "24", "", ""],
    ["Melarsoprol", "Tripanossomíase africana", "pediatrica", "IV", "peso(1, 40); Idade(30, 6570) dosagem(2.2, 3.6, 2.2, mg/kg).", "36", "24", "", ""],
    ["Eflornitina", "Tripanossomíase africana", "adulta", "IV", "peso(40, 150); Idade(6570, 36500) dosagem(100, 100, 100, mg/kg).", "200", "6", "", ""],
    ["Eflornitina", "Tripanossomíase africana", "pediatrica", "IV", "peso(1, 40); Idade(30, 6570) dosagem(100, 100, 100, mg/kg).", "200", "6", "", ""],
    ["Quinina", "Malária grave (Ataque)", "adulta", "IV", "peso(40, 150); Idade(4381, 36500) dosagem(20, 20, 20, mg/kg).", "300", "", "", ""],
    ["Quinina", "Malária grave (Ataque)", "pediatrica", "IV", "peso(1, 40); Idade(30, 4380) dosagem(20, 20, 20, mg/kg).", "300", "", "", ""],
    ["Quinina", "Malária grave (Manutenção)", "adulta", "IV", "peso(40, 150); Idade(4381, 36500) dosagem(10, 10, 10, mg/kg).", "300", "8", "", ""],
    ["Quinina", "Malária grave (Manutenção)", "pediatrica", "IV", "peso(1, 40); Idade(30, 4380) dosagem(10, 10, 10, mg/kg).", "300", "8", "", ""],
    ["Quinina", "Malária grave (Ataque)", "adulta", "IM", "peso(40, 150); Idade(4381, 36500) dosagem(10, 10, 10, mg/kg).", "300", "4", "", ""],
    ["Quinina", "Malária grave (Ataque)", "pediatrica", "IM", "peso(1, 40); Idade(30, 4380) dosagem(10, 10, 10, mg/kg).", "300", "4", "", ""],
    ["Artesunato", "Malária grave (Dia 1)", "adulta", "IV", "peso(40, 150); Idade(4381, 36500) dosagem(2.4, 2.4, 2.4, mg/kg).", "60", "12", "", ""],
    ["Artesunato", "Malária grave (Dia 1)", "pediatrica", "IV", "peso(1, 40); Idade(30, 4380) dosagem(2.4, 3, 2.4, mg/kg).", "60", "12", "", ""],
    ["Artesunato", "Malária grave (Manutenção)", "adulta", "IV", "peso(40, 150); Idade(4381, 36500) dosagem(2.4, 2.4, 2.4, mg/kg).", "60", "24", "", ""],
    ["Artesunato", "Malária grave (Manutenção)", "pediatrica", "IV", "peso(1, 40); Idade(30, 4380) dosagem(2.4, 3, 2.4, mg/kg).", "60", "24", "", ""],
    ["Artesunato", "Malária grave (Dia 1)", "adulta", "IM", "peso(40, 150); Idade(4381, 36500) dosagem(2.4, 2.4, 2.4, mg/kg).", "60", "12", "", ""],
    ["Artesunato", "Malária grave (Dia 1)", "pediatrica", "IM", "peso(1, 40); Idade(30, 4380) dosagem(2.4, 3, 2.4, mg/kg).", "60", "12", "", ""],
    ["Artesunato", "Malária grave (Manutenção)", "adulta", "IM", "peso(40, 150); Idade(4381, 36500) dosagem(2.4, 2.4, 2.4, mg/kg).", "60", "24", "", ""],
    ["Artesunato", "Malária grave (Manutenção)", "pediatrica", "IM", "peso(1, 40); Idade(30, 4380) dosagem(2.4, 3, 2.4, mg/kg).", "60", "24", "", ""],
    ["Artemeter", "Malária grave (Ataque)", "adulta", "IM", "peso(40, 150); Idade(4381, 36500) dosagem(3.2, 3.2, 3.2, mg/kg).", "80", "", "", ""],
    ["Artemeter", "Malária grave (Ataque)", "pediatrica", "IM", "peso(1, 40); Idade(30, 4380) dosagem(3.2, 3.2, 3.2, mg/kg).", "80", "", "", ""],
    ["Artemeter", "Malária grave (Manutenção)", "adulta", "IM", "peso(40, 150); Idade(4381, 36500) dosagem(1.6, 1.6, 1.6, mg/kg).", "80", "24", "", ""],
    ["Artemeter", "Malária grave (Manutenção)", "pediatrica", "IM", "peso(1, 40); Idade(30, 4380) dosagem(1.6, 1.6, 1.6, mg/kg).", "80", "24", "", ""],
    ["Benznidazol", "Doença de Chagas", "adulta", "IV", "peso(40, 150); Idade(6570, 36500) dosagem(5, 7, 5, mg/kg).", "100", "12", "", ""],
    ["Benznidazol", "Doença de Chagas", "pediatrica", "IV", "peso(1, 40); Idade(0, 6570) dosagem(5, 10, 5, mg/kg).", "100", "12", "", ""],
    ["Nifurtimox", "Doença de Chagas", "pediatrica", "IV", "peso(1, 40); Idade(30, 6570) dosagem(10, 20, 15, mg/kg).", "120", "8, 12", "", ""],
    ["Rifampicina", "Hanseníase", "adulta", "IV", "peso(40, 150); Idade(6570, 36500) dosagem(600, 600, 600, mg).", "600", "24", "", ""],
    ["Rifampicina", "Hanseníase", "pediatrica", "IV", "peso(1, 40); Idade(30, 6570) dosagem(10, 10, 10, mg/kg).", "600", "24", "", ""],
    ["Clindamicina", "Malária (com Quinino)", "adulta", "IV", "peso(40, 150); Idade(4381, 36500) dosagem(10, 10, 10, mg/kg).", "150", "8", "", ""],
    ["Clindamicina", "Malária (com Quinino)", "pediatrica", "IV", "peso(1, 40); Idade(30, 4380) dosagem(10, 10, 10, mg/kg).", "150", "8", "", ""],
    ["Dapsona", "Profilaxia PCP (Alérgia)", "adulta", "IV", "peso(40, 150); Idade(4381, 36500) dosagem(100, 100, 100, mg).", "100", "24", "", ""],
    ["Dapsona", "Profilaxia PCP (Alérgia)", "pediatrica", "IV", "peso(1, 40); Idade(30, 4380) dosagem(2, 2, 2, mg/kg).", "100", "24", "", ""],
    ["Ansuvimab", "Ébola", "adulta", "IV", "peso(40, 150); Idade(4381, 36500) dosagem(50, 50, 50, mg/kg).", "400", "", "", ""],
    ["Ansuvimab", "Ébola", "pediatrica", "IV", "peso(1, 40); Idade(0, 4380) dosagem(50, 50, 50, mg/kg).", "400", "", "", ""],
    ["Atoltivimab + Maftivimab + Odesivimab", "Ébola", "adulta", "IV", "peso(40, 150); Idade(4381, 36500) dosagem(50, 50, 50, mg/kg).", "150", "", "", ""],
    ["Atoltivimab + Maftivimab + Odesivimab", "Ébola", "pediatrica", "IV", "peso(1, 40); Idade(0, 4380) dosagem(50, 50, 50, mg/kg).", "150", "", "", ""],
    ["Ivermectina", "Estrongiloidíase grave", "adulta", "SC", "peso(40, 150); Idade(6570, 36500) dosagem(200, 200, 200, mcg/kg).", "10", "24", "", ""],
    ["Ivermectina", "Estrongiloidíase grave", "pediatrica", "SC", "peso(15, 40); Idade(1825, 6570) dosagem(200, 200, 200, mcg/kg).", "10", "24", "", ""],
    ["Praziquantel", "Esquistossomose grave", "adulta", "IV", "peso(40, 150); Idade(6570, 36500) dosagem(40, 60, 40, mg/kg).", "600", "", "", ""],
    ["Praziquantel", "Esquistossomose grave", "pediatrica", "IV", "peso(1, 40); Idade(730, 6570) dosagem(40, 60, 40, mg/kg).", "600", "", "", ""],
    ["Albendazol", "Neurocisticercose grave", "adulta", "IV", "peso(40, 150); Idade(6570, 36500) dosagem(15, 15, 15, mg/kg).", "200", "12", "", ""],
    ["Albendazol", "Neurocisticercose grave", "pediatrica", "IV", "peso(1, 40); Idade(730, 6570) dosagem(15, 15, 15, mg/kg).", "200", "12", "", ""],
    ["Diloxanida", "Amebíase extraintestinal", "adulta", "IV", "peso(40, 150); Idade(6570, 36500) dosagem(500, 500, 500, mg).", "500", "8", "", ""],
    ["Diloxanida", "Amebíase extraintestinal", "pediatrica", "IV", "peso(10, 40); Idade(730, 6570) dosagem(20, 20, 20, mg/kg).", "500", "8", "", ""],
    ["Tidafusina", "Hepatite B", "adulta", "SC", "peso(40, 150); Idade(6570, 36500) dosagem(1.6, 1.6, 1.6, mg).", "1.6", "48, 72", "", ""],
    ["Zidovudina", "Profilaxia HIV (RN)", "pediatrica", "IV", "peso(1, 5); Idade(0, 30) dosagem(1.5, 1.5, 1.5, mg/kg).", "10", "6, 12", "", ""],
    ["Nevirapina", "Profilaxia HIV (RN)", "pediatrica", "IV", "peso(1, 5); Idade(0, 30) dosagem(2, 2, 2, mg/kg).", "10", "24", "", ""],
    ["Amikacina", "Meningite por Gram-negativos", "adulta", "Intratecal", "peso(40, 150); Idade(4381, 36500) dosagem(5, 50, 20, mg).", "250", "24", "", ""],
    ["Amikacina", "Meningite por Gram-negativos", "pediatrica", "Intratecal", "peso(1, 40); Idade(0, 4380) dosagem(1, 5, 2, mg).", "250", "24", "", ""],
    ["Gentamicina", "Meningite por Gram-negativos", "adulta", "Intratecal", "peso(40, 150); Idade(4381, 36500) dosagem(1, 5, 4, mg).", "40", "24", "", ""],
    ["Gentamicina", "Meningite por Gram-negativos", "pediatrica", "Intratecal", "peso(1, 40); Idade(0, 4380) dosagem(1, 2, 1, mg).", "40", "24", "", ""],
    ["Vancomicina", "Meningite por MRSA", "adulta", "Intratecal", "peso(40, 150); Idade(4381, 36500) dosagem(5, 20, 10, mg).", "500", "24", "", ""],
    ["Vancomicina", "Meningite por MRSA", "pediatrica", "Intratecal", "peso(1, 40); Idade(0, 4380) dosagem(2, 5, 3, mg).", "500", "24", "", ""],
    ["Colistina", "Meningite MDR", "adulta", "Intratecal", "peso(40, 150); Idade(4381, 36500) dosagem(125000, 125000, 125000, UI).", "80", "24", "", ""],
    ["Colistina", "Meningite MDR", "pediatrica", "Intratecal", "peso(1, 40); Idade(0, 4380) dosagem(2000, 2000, 2000, UI/kg).", "80", "24", "", ""],
    ["Polimixina B", "Meningite MDR", "adulta", "Intratecal", "peso(40, 150); Idade(4381, 36500) dosagem(50000, 50000, 50000, UI).", "500000 UI|500000", "24", "", ""],
    ["Polimixina B", "Meningite MDR", "pediatrica", "Intratecal", "peso(1, 40); Idade(0, 4380) dosagem(10000, 20000, 20000, UI).", "500000 UI|500000", "24", "", ""],
    ["Surfactante (Poractante)", "SDR neonatal", "pediatrica", "Intratraqueal", "peso(0.5, 5); Idade(0, 7) dosagem(100, 200, 200, mg/kg).", "80", "6, 12", "", ""],
    ["Dinitrato de Isossorbida", "Angina", "adulta", "IV", "peso(40, 150); Idade(4381, 36500) dosagem(2, 12, 5, mg/h).", "1", "", "", ""],
    ["Nitroprussiato de Sodio", "Emergência hipertensiva", "adulta", "IV", "peso(40, 150); Idade(4381, 36500) dosagem(0.3, 10, 3, mcg/kg/min).", "10", "", "", ""],
    ["Nitroprussiato de Sodio", "Emergência hipertensiva", "pediatrica", "IV", "peso(1, 40); Idade(30, 4380) dosagem(0.3, 8, 3, mcg/kg/min).", "10", "", "", ""],
    ["Fenoldopam", "Emergência hipertensiva", "adulta", "IV", "peso(40, 150); Idade(4381, 36500) dosagem(0.1, 1.6, 0.1, mcg/kg/min).", "10", "", "", ""],
    ["Epoprostenol", "Hipertensão pulmonar", "adulta", "IV", "peso(40, 150); Idade(4381, 36500) dosagem(2, 40, 2, ng/kg/min).", "0.5, 1.5", "", "", ""],
    ["Treprostinilo", "Hipertensão pulmonar", "adulta", "IV", "peso(40, 150); Idade(4381, 36500) dosagem(1.25, 40, 1.25, ng/kg/min).", "1, 2.5, 5, 10", "", "", ""],
    ["Fisostigmina", "Reversão anticolinérgica", "adulta", "IV", "peso(40, 150); Idade(4381, 36500) dosagem(0.5, 2, 1, mg).", "1", "0.33", "", ""],
    ["Fisostigmina", "Reversão anticolinérgica", "pediatrica", "IV", "peso(1, 40); Idade(30, 4380) dosagem(0.02, 0.02, 0.02, mg/kg).", "1", "0.33", "", ""],
    ["Glicopirrolato", "Reversão bloqueio muscular", "adulta", "IV", "peso(40, 150); Idade(4381, 36500) dosagem(0.2, 0.2, 0.2, mg).", "0.2", "", "", ""],
    ["Glicopirrolato", "Reversão bloqueio muscular", "pediatrica", "IV", "peso(1, 40); Idade(30, 4380) dosagem(0.01, 0.01, 0.01, mg/kg).", "0.2", "", "", ""],
    ["Edrofonio", "Teste diagnóstico Miastenia", "adulta", "IV", "peso(40, 150); Idade(4381, 36500) dosagem(2, 10, 2, mg).", "10", "", "", ""],
    ["Edrofonio", "Teste diagnóstico Miastenia", "pediatrica", "IV", "peso(1, 40); Idade(30, 4380) dosagem(0.1, 0.2, 0.2, mg/kg).", "10", "", "", ""],
    ["Verapamil", "Crise hipertensiva", "adulta", "IV", "peso(40, 150); Idade(4381, 36500) dosagem(5, 10, 5, mg).", "2.5", "0.25", "", ""],
    ["Diltiazem", "Hipertensão grave", "adulta", "IV", "peso(40, 150); Idade(4381, 36500) dosagem(5, 15, 10, mg/h).", "5", "", "", ""],
    ["Esmolol", "Taquicardia supraventricular", "adulta", "IV", "peso(40, 150); Idade(4381, 36500) dosagem(50, 300, 100, mcg/kg/min).", "10, 250", "", "", ""],
    ["Esmolol", "Taquicardia supraventricular", "pediatrica", "IV", "peso(1, 40); Idade(30, 4380) dosagem(50, 500, 200, mcg/kg/min).", "10, 250", "", "", ""],
    ["Metoprolol", "Crise adrenérgica", "adulta", "IV", "peso(40, 150); Idade(4381, 36500) dosagem(1, 5, 1, mg).", "1", "0.08", "", ""],
    ["Phentolamina", "Crise feocromocitoma", "adulta", "IV", "peso(40, 150); Idade(4381, 36500) dosagem(2, 5, 5, mg).", "5", "", "", ""],
    ["Phentolamina", "Crise feocromocitoma", "pediatrica", "IV", "peso(1, 40); Idade(30, 4380) dosagem(0.05, 0.1, 0.1, mg/kg).", "5", "", "", ""],
    ["Argatroban", "Trombocitopenia (HIT)", "adulta", "IV", "peso(40, 150); Idade(4381, 36500) dosagem(2, 10, 2, mcg/kg/min).", "1", "", "", ""],
    ["Bivalirudina", "Angioplastia coronária", "adulta", "IV", "peso(40, 150); Idade(4381, 36500) dosagem(0.75, 1.75, 0.75, mg/kg).", "250", "", "", ""],
    ["Sufentanilo", "Anestesia equilibrada", "adulta", "IV", "peso(40, 150); Idade(4381, 36500) dosagem(0.5, 5, 1, mcg/kg).", "5 mcg/mL|5, 50 mcg/mL|50", "", "", ""],
    ["Sufentanilo", "Anestesia equilibrada", "pediatrica", "IV", "peso(1, 40); Idade(30, 4380) dosagem(0.1, 0.5, 0.2, mcg/kg).", "5 mcg/mL|5, 50 mcg/mL|50", "", "", ""],
    ["Alfentanilo", "Indução rápida", "adulta", "IV", "peso(40, 150); Idade(4381, 36500) dosagem(8, 50, 20, mcg/kg).", "0.5", "", "", ""],
    ["Alfentanilo", "Indução rápida", "pediatrica", "IV", "peso(1, 40); Idade(30, 4380) dosagem(10, 20, 15, mcg/kg).", "0.5", "", "", ""],
    ["Sumatriptan", "Cefaleia vascular", "adulta", "SC", "peso(40, 150); Idade(6570, 36500) dosagem(6, 6, 6, mg).", "12", "12", "", ""]
]



motor_horizontal('medicamentos.xlsx', novos_dados)